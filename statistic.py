import csv
import matplotlib
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, numbers
import numpy as np
import matplotlib.pyplot as plt
from jinja2 import Environment, FileSystemLoader
import pdfkit
import base64


class DataSet:
    """Класс для формирования статистики

    Attributes:
        file_name (str): Название csv-файла, из которого берутся данные
        jobName = Профессия по которой собирается статистика
        keys = Столбцы csv-файла
        year_salary = Статистика оклада по годам
        year_count = Статистика количества вакансий по годам
        job_year_salary = Статистика оклада по годам для выбранной профессии
        job_year_count = Статистика количества вакансий по годам для выбранной профессии
        city_salary = Статистика оклада по городам
        city_precent = Статистика вакансий по городам
    """
    def __init__(self, _file_name, _job):
        """Инициализирует объект DataSet

        :param _file_name: Файл из которого берутся данные
        :param _job: Выбранная профессия
        """
        self.file_name = _file_name
        self.jobName = _job
        self.keys = {}

        self.year_salary = {}
        self.year_count = {}
        self.job_year_salary = {}
        self.job_year_count = {}
        self.city_salary = {}
        self.city_precent = {}
        self.total_city = 0

        self.csv_reader(self.file_name)

    def csv_reader(self, file_name):
        """Считывает данные из csv-файла и формирует статистику

        :param file_name: csv-файл
        """
        with open(file_name, newline='', encoding='utf-8-sig') as File:
            reader = csv.reader(File, delimiter=',')
            for row in reader:
                if not self.keys:
                    self.keys['name'] = row.index('name')
                    self.keys['salary_from'] = row.index('salary_from')
                    self.keys['salary_to'] = row.index('salary_to')
                    self.keys['salary_currency'] = row.index('salary_currency')
                    self.keys['area_name'] = row.index('area_name')
                    self.keys['published_at'] = row.index('published_at')
                    continue
                if row.__contains__(""):
                    continue
                self.add_year_statistic(row)
                self.add_city_statistic(row)
        self.aver_salary()
        self.calculate_city_precent()
        self.aver_job_salary()

    def add_year_statistic(self, row):
        """Формирует статистику по годам

        :param row: csv-строка
        """
        year = int(row[self.keys['published_at']][:4])
        if year not in self.year_salary.keys():
            self.year_salary[year] = 0
            self.year_count[year] = 0
            self.job_year_salary[year] = 0
            self.job_year_count[year] = 0
        self.year_salary[year] += self.calculate_salary(row)
        self.year_count[year] += 1

        if jobName in row[self.keys['name']]:
            self.job_year_salary[year] += self.calculate_salary(row)
            self.job_year_count[year] += 1

    def add_city_statistic(self, row):
        """Формирует статистику по городам

        :param row: csv-строка
        """
        city = row[self.keys['area_name']]
        if city not in self.city_salary.keys():
            self.city_salary[city] = 0
            self.city_precent[city] = 0
        self.total_city += 1
        self.city_salary[city] += self.calculate_salary(row)
        self.city_precent[city] += 1

    def calculate_salary(self, row):
        """Считает среднюю зарплату для вакансии

        :param row: csv-строка
        :return: Средняя зарплата
        """
        from_ = float(row[self.keys['salary_from']]) * self.currency_to_rub[row[self.keys['salary_currency']]]
        to_ = float(row[self.keys['salary_to']]) * self.currency_to_rub[row[self.keys['salary_currency']]]
        return (from_ + to_) / 2

    def aver_salary(self):
        """Считает среднюю зарплату для каждого года

        """
        for year in self.year_count:
            self.year_salary[year] = int(self.year_salary[year] / self.year_count[year])

    def aver_job_salary(self):
        """Считает среднюю зарплату для каждого года для выбранной профессии

        """
        for year in self.job_year_salary:
            if not self.job_year_count[year] == 0:
                self.job_year_salary[year] = int(self.job_year_salary[year] / self.job_year_count[year])

    def calculate_city_precent(self):
        """Считает процент вакансий для городов,
        оставляет только с процетом больше 1

        """
        cities = {y: round(self.city_precent[y]/self.total_city, 4) for y in self.city_precent.keys()
                  if self.city_precent[y]/self.total_city >= 0.01}
        self.calculate_city_salary(cities)
        self.city_precent = dict(sorted(cities.items(),  key=lambda item: item[1], reverse=True)[:10])

    def calculate_city_salary(self, cities):
        """Считает среднюю зарплату для городов,
        процент которых больше 1

        """
        salary = {}

        for city in cities.keys():
            salary[city] = int(self.city_salary[city] / self.city_precent[city])

        self.city_salary = dict(sorted(salary.items(), key=lambda item: item[1], reverse=True)[:10])

    def print_data(self):
        """Выводит статистику в консоль

        """
        print(f"Динамика уровня зарплат по годам: {self.year_salary}")
        print(f"Динамика количества вакансий по годам: {self.year_count}")
        print(f"Динамика уровня зарплат по годам для выбранной профессии: {self.job_year_salary}")
        print(f"Динамика количества вакансий по годам для выбранной профессии: {self.job_year_count}")
        print(f"Уровень зарплат по городам (в порядке убывания): {self.city_salary}")
        print(f"Доля вакансий по городам (в порядке убывания): {self.city_precent}")

    currency_to_rub = {
        "AZN": 35.68,
        "BYR": 23.91,
        "EUR": 59.90,
        "GEL": 21.74,
        "KGS": 0.76,
        "KZT": 0.13,
        "RUR": 1,
        "UAH": 1.64,
        "USD": 60.66,
        "UZS": 0.0055,
    }


class Report:
    """Класс для формирования отчётов со статистикой

        jobName = Профессия по которой собирается статистика
        year_salary = Статистика оклада по годам
        year_count = Статистика количества вакансий по годам
        job_year_salary = Статистика оклада по годам для выбранной профессии
        job_year_count = Статистика количества вакансий по годам для выбранной профессии
        city_salary = Статистика оклада по городам
        city_precent = Статистика вакансий по городам
    """
    def __init__(self, data):
        """Инициализирует объект Report

        :param data: объект DataSet
        """
        self.year_salary = data.year_salary
        self.job_year_salary = data.job_year_salary
        self.year_count = data.year_count
        self.job_year_count = data.job_year_count
        self.job_name = data.jobName
        self.city_salary = data.city_salary
        self.city_precent = data.city_precent

        self.border = Border(
            left=Side(border_style="medium", color='000000'),
            right=Side(border_style="medium", color='000000'),
            top=Side(border_style="medium", color='000000'),
            bottom=Side(border_style="medium", color='000000'))

    def generate_excel(self):
        """Формирует exel-таблицу со статистикой
        и сохраняет в текущей деректории

        """
        wb = Workbook()
        year_ws = wb.active
        year_ws.title = "Статистика по годам"
        year_ws.append(['Год', 'Средняя зарплата', f'Средняя зарплата - {self.job_name}', 'Количество вакансий',
                        f'Количество вакансий - {self.job_name}'])
        for year in self.year_salary.keys():
            year_ws.append([year, self.year_salary[year], self.job_year_salary[year],
                            self.year_count[year], self.job_year_count[year]])
        self.stylised(year_ws)

        city_ws = wb.create_sheet("Статистика по городам")
        city_ws.append(['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий'])
        city1 = list(self.city_salary.keys())
        city2 = list(self.city_precent.keys())
        for row in range(2, 12):
            for col in range(1, 6):
                if col == 1:
                    city_ws.cell(row=row, column=col, value=city1[row-2])
                elif col == 2:
                    city_ws.cell(row=row, column=col, value=self.city_salary[city1[row-2]])
                elif col == 3:
                    city_ws.cell(row=row, column=col, value='')
                elif col == 4:
                    city_ws.cell(row=row, column=col, value=city2[row-2])
                elif col == 5:
                    city_ws.cell(row=row, column=col, value=self.city_precent[city2[row-2]])\
                        .number_format = numbers.BUILTIN_FORMATS[10]

        self.stylised(city_ws)

        wb.save('test.xlsx')

    def stylised(self, ws):
        """Стилизует таблицу

        :param ws: exel-таблица
        """
        column_widths = []
        for row in ws:
            for i, cell in enumerate(row):
                if len(column_widths) > i:
                    if len(str(cell.value)) > column_widths[i]:
                        column_widths[i] = len(str(cell.value))+2
                else:
                    column_widths += [len(cell.value)+2]
                if not cell.value == '':
                    cell.border = self.border

        for i, column_width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = column_width

    def generate_image(self):
        """Формирует изображения с 4 графиками статистики

        """
        matplotlib.use('TkAgg')
        plt.rc("font", size=8)

        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2)
        fig.set_size_inches(8,6)
        self.make_bar_chart(ax1, "Уровень зарплат по годам", "средняя з/п", f"з/п {self.job_name}",
                            self.year_salary, self.job_year_salary)
        self.make_bar_chart(ax2, "Количество вакансий по годам", "количество вакансий",
                            f"количество вакансий {self.job_name}", self.year_count, self.job_year_count)
        self.make_horizontal_chart(ax3, "Уровень зарплат по городам", self.city_salary)
        self.make_pie_chart(ax4, 'Доля вакансий по городам', self.city_precent)


        fig.tight_layout()

        plt.savefig("graph.png")
        plt.show()

    def make_bar_chart(self, ax,  title, bar1, bar2, dict1, dict2):
        """Формирует гистограмму

        :param ax: область где отображается график
        :param title: Название гистограммы
        :param bar1: Название первого столбца в легенде
        :param bar2: Название второго столбца в легенде
        :param dict1: данные для первого столбца
        :param dict2: данные для второго столбца
        """
        x = np.arange(len(list(dict1.keys())))
        width = 0.35

        ax.bar(x - width/2, [dict1[x] for x in dict1], width, label=bar1, tick_label=None)
        ax.bar(x + width/2, [dict2[x] for x in dict2], width, label=bar2, tick_label=None)

        ax.set_title(title)
        ax.grid(axis='y')
        ax.set_xticks(x, list(dict1.keys()), rotation=90, fontsize=8)
        ax.legend(fontsize=8)

    def make_horizontal_chart(self, ax,  title, dict1):
        """Формирует горизонтальную гистограмму

        :param ax: Область где отображается график
        :param title: Название графика
        :param dict1: Данные для графика
        """
        newdict = []
        for city in reversed(dict1):
            if " " in city:
                newdict.append(city.replace(" ", "\n"))
            elif "-" in city:
                newdict.append(city.replace('-', '-\n'))
            else:
                newdict.append(city)

        y = np.arange(len(newdict))
        ax.barh(y, [dict1[x] for x in reversed(dict1)], align='center')

        ax.set_yticks(y, labels=newdict, fontsize=8)
        ax.set_title(title)
        ax.grid(axis='x')

    def make_pie_chart(self, ax, title, data):
        """Формирует круговую диаграмму

        :param ax: Область где отображается график
        :param title: Название графика
        :param data: Данные для графика
        """
        precents = list(reversed(data.values()))
        precents.append(1 - sum(precents))

        labels = list(reversed(data.keys()))
        labels.append('Другое')

        ax.pie(precents, labels=labels, textprops={'fontsize': 6})

        ax.set_title(title)

    def generate_pdf(self):
        """Формирует pdf-отчёт с таблицами и графиками
        со статистикой. Сохраняет его в текущей директории

        """
        env = Environment(loader=FileSystemLoader('.'))
        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')

        items1 = []
        items2 = []

        for i in self.year_salary:
            items1.append(dict(year=i, salary=self.year_salary[i], jobsalary=self.job_year_salary[i],
                               count=self.year_count[i], jobcount=self.job_year_count[i]))

        ck1 = list(self.city_salary.keys())
        ck2 = list(self.city_precent.keys())
        for i in range(len(ck2)):
            c1 = ck1[i]
            c2 = ck2[i]

            items2.append(dict(city1=c1, salary=self.city_salary[c1], city2=c2,
                               precent=self.preform(self.city_precent[c2])))

        template = env.get_template("report_template.html")
        pdf_template = template.render(items1=items1, items2=items2, job=self.job_name,
                                       img_file=self.image_converter("graph.png"))

        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config)

    def image_converter(self, filepath):
        """Конвертирует путь до изображения в base64

        :param filepath: путь до изображения
        :return: путь в формате base64
        """
        with open(filepath, 'rb') as f:
            return base64.b64encode(f.read()).decode()

    def preform(self, digit):
        """Прифодит проценты к нужному формату

        :param digit: число
        :return: строка процетов в нужном формате
        """
        return "{0:.2f}".format(digit*100) + "%"



fileName_ = input("Введите название файла: ")
jobName = input("Введите название профессии: ")

statistic = DataSet(fileName_, jobName)


rp = Report(statistic)
rp.generate_image()
rp.generate_pdf()
